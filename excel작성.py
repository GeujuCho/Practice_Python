#브라우저 실행
from selenium import webdriver
driver=webdriver.Chrome('C:/Users/soldesk/bigdata/chromedriver.exe')

#상장회사 검색
driver.get('http://marketdata.krx.co.kr/mdi#document=040601')

from selenium.webdriver.common.by import By
button=driver.find_element(By.XPATH, '//button[text()="Excel"]')
button.click()

import os
import time
folder='C:/Users/soldesk/Downloads'
os.chdir(folder)

fname='data.xls'
for _ in range(30):
    if os.path.exists(fname):
        break
        time.sleep(1)

folder='C:/Users/soldesk/bigdata'
os.chdir(folder)

# os.rename('data.xls','상장회사목록.xls')
os.getcwd()

import pandas as pd
df_삼성전자=pd.read_excel('2017년 광고 삼성전자.xlsx')
df_삼성전자.set_index('date',inplace=True)
df_LG전자=pd.read_excel('2017년 광고 엘지전자.xlsx')
df_LG전자.set_index('date',inplace=True)

df_merge=pd.DataFrame()
df_merge['삼성전자']=df_삼성전자['total']
df_merge['LG전자']=df_LG전자['total']

df_merge.to_excel('merged_01.xlsx')

import openpyxl

wb=openpyxl.load_workbook('merged_01.xlsx')
sheet=wb.active
sheet

삼성전자_월광고비=[row[0].value for row in sheet['B2':'B13']]
삼성전자_월광고비

sum(삼성전자_월광고비)

sheet['A28'].value='합계'

wb.save('merged_01.xlsx')

sheet['B28'].value='=SUM(B2:B13)'
sheet['C28'].value='=SUM(C2:C13)'

wb.save('merged_01.xlsx')

from openpyxl.styles import Font, Alignment, Border,Side ,Color,PatternFill

font_15=Font(name='맑은 고딕',size=15,bold=True)
align_center=Alignment(horizontal='center',vertical='center')
align_vcenter=Alignment(vertical='center')

border_thin=Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'),bottom=Side(style='thin'))

fill_orange=PatternFill(patternType='solid', fgColor=Color('ffc000'))
fill_lightgray=PatternFill(patternType='solid', fgColor=Color('d3d3d3'))

cell_sum=sheet['A28']

cell_sum.font=font_15
cell_sum.alignment=align_center
cell_sum.border=border_thin
cell_sum.fill=fill_orange

wb.save('merged_01.xlsx')

for row in sheet['B2:C14']:
    for cell in row:
        cell.border=border_thin
        cell.number_format='0.00'

for row in sheet['B14:C14']:
    for cell in row:
        cell.alignment=align_vcenter
        cell.fill=fill_orange

# 셀의 크기 늘리기
for col in sheet['A2:C28']:
     max_length = 0
     for cell in range(3):
        if(cell%2==0):
            col[cell].font=font_15
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                 pass
     adjusted_width = (max_length + 2) * 10
for i in range(0, 3):
        sheet.column_dimensions[col[i].column].width = adjusted_width

wb.save('merged_01.xlsx')
