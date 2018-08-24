import openpyxl

# 엑셀 파일 가져오기
wb=openpyxl.load_workbook('c.xlsx')
ws=wb.active
ws

# 빈 엑셀 파일 만들기
wbN=openpyxl.Workbook() #새로운 파일 열기
wsN=wbN.active  #열린 워크북의 처음 worksheet활성화
wsN

# 데이터 복사하기
for select in ws.iter_rows("A1:C29"):
    for i in select:
        print(i.coordinate)
        wsN[i.coordinate].value=i.value  # coordinate를 이용하여 엑셀의 좌표를 정해줘야 한다.

# 새로운  파일명으로 저장하기
wbN.save('exam1.xlsx')