import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font

# 엑셀 파일 불러오기
wb=openpyxl.load_workbook('c.xlsx')
sheet=wb.active
sheet

# 바꿀 폰트의 특징 정하기
font_1=Font(name='돋음',size=20,bold=True)

# iter_rows를 이용하여 바꾸기
for row in sheet.iter_rows("A1:C27"):
    for cell in row:
        if cell.value==1:
            sheet[str(cell.coordinate)].font = font_1
            print(cell.coordinate)

wb.save('test01.xlsx')

# iter_cols를 이용하여 바꾸기
for col in sheet.iter_cols(min_row=0,min_col=0,max_row=29,max_col=3):
    for i in range(29):
        if(col[i].value==1):
            print(col[i].coordinate)
            col[i].font=font_1

wb.save('test02.xlsx')